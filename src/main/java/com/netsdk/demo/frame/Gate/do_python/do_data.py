# * coding:utf-8 *
"""
 # File   : do_data.PY
 # Time   : 2021/4/6 16:02
 # Author : Lethe
 # version: python 3.8
"""
from builtins import print as _print
from sys import _getframe
from sqlalchemy import create_engine, MetaData, Table, update, select, text
from datetime import datetime
from sqlalchemy.sql import func

import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from apscheduler.schedulers.background import BackgroundScheduler
import openpyxl as ex


def print(*arg, **kw):
    s = f'{_getframe(1).f_lineno}行：'
    return _print(s, *arg, **kw)


# 初始化
# engine = create_engine('sqlite:///..\\gate.db', future=True)
engine = create_engine('sqlite:///..\\gate.db', future=True, connect_args={'check_same_thread': False})
metadata = MetaData()
in_table = Table('InRecode', metadata, autoload=True, autoload_with=engine)
out_table = Table('OutRecode', metadata, autoload=True, autoload_with=engine)
con = engine.connect()
# session = Session(engine)

# with engine.connect() as con:
#     # rs = con.execute('select * from InRecode')
#     rs = con.execute(text("select 'hello world'"))
#     print(rs.fetchone())

# # 使用begin提交数据，不用再调用commit()方法。
# with engine.begin() as conn:
#     res = conn.execute(text("INSERT INTO InRecode (name, nameid, datetime, state) VALUES (:name, :nameid, :datetime, :state)"),
#                  [{"name": "韩梅梅", "nameid": 123456, "datetime": "2021-4-1 18:00:00", "state": 1}])

# # 读取查询的数据
# with engine.begin() as conn:
#     res = conn.execute(text("select * from InRecode"))
#     for i in res:
#         print(i.Name, i.NameId, i.DATETIME, i.STATE)


# # 使用Session插入数据或查询，用 bindparams 绑定参数。
# sentence = "INSERT INTO InRecode (name, nameid, datetime) VALUES (:name, :nameid, :datetime)"
# stmt = text(sentence).bindparams(name='Marry', nameid=111111, datetime='2021-04-08 8:8:8')
# with Session(engine) as session:
#     session.execute(stmt)
#     session.commit()

# print(in_table)

# 查询数据
# res = session.query(in_table).all()
# print(res)

# 查询出门的记录
def query_out(name_id):
    # ☆☆☆ 暂时安名字进行查询，之后需把 out_table.c.Name 改为 out_table.c.NameId
    stmt_out = select(out_table).where(out_table.c.NameId == name_id)
    res_out = con.execute(stmt_out)
    # print(dir(res_out))
    res_list = res_out.all()

    if res_list:    # 获取最后一条出门记录
        # print(dir(res_list[-1]))
        # print(res_list[-1].Name, res_list[-1].DATETIME)
        return res_list[-1]
    else:
        # print("无出门数据！")
        return None


# select 方法查询
# 查询已进门但未出门的记录
def query_in():
    stmt_in = select(in_table).where(in_table.c.STATE == 0)
    res_in = con.execute(stmt_in)
    in_list = res_in.all()
    # print(in_list)
    return in_list


# 综合进出记录，筛选出在房内人员
def read_indoor_people():
    in_list = query_in()
    # print(len(in_list))

    # 生成当前未出门的字典数据，剔除了个人的重复记录。
    in_dict = {}
    for i in in_list:
        # print(i)
        in_dict.update({i.NameId: i})  # ☆☆☆ 字典键暂时为人员 名字， 今后改为 i.NameId !!!!!!!
    # print(in_dict)

    for i in in_dict.copy():           # i 为 in_dict 字典的 key。
        n = query_out(i)               # 查询人员最后一次出门记录，对in_dict字典进行清洗。
        if n:
            try:
                out_time = n.DATETIME
                in_time = in_dict.get(i).DATETIME
                if (out_time - in_time).total_seconds() > 0:
                    # in_dict.get(i).STATE = 1
                    # 对已出门的人员的状态进行更新。
                    con.execute(update(in_table).where(in_table.c.NameId == i, in_table.c.STATE == 0).values(STATE=1))
                    con.commit()
                    in_dict.pop(i)  # 剔除已出门的人员
            except Exception as e:
                con.execute(update(in_table).where(in_table.c.NameId == i, in_table.c.STATE == 0).values(STATE=-1))
                con.commit()
                print(e)

    # print("在内部人员：", in_dict)  # 生成最终未出门的人员。
    # 以字典格式返回，如：{'小红的ID': (9, '小红', '331022198610013119', datetime.datetime(2021, 4, 8, 14, 31, 48), 0)}
    return in_dict


# 定义一个显示的函数，返回一个在房间内人员的字典。
def show_people_info():
    global SHOW
    SHOW = False
    people = read_indoor_people()
    people_num = len(people)
    v.set(f"当前人数为：{people_num}人")
    # 渲染到 text 窗口
    # 参数 people 为 read_indoor_people() 读取的数据字典。
    text.config(state='normal')
    text.delete(1.0, "end")
    # 第一次渲染人员信息
    text.insert('end', '|序号|   姓名   |         ID号         |        进门时间         |\n')
    text.insert('end', '—' * 33 + "\n")
    for n, i in enumerate(people, start=1):
        # print(people[i])
        # print(len(f"{people[i]['Name']:^7}"))
        n_ord = 0
        for s in people[i]['Name']:
            if ord(s) > 255:
                n_ord += 1                               # 获取名字中的汉字数，一个汉字占两个字节。
        sup_len = 10 - (len(people[i]['Name'])+n_ord)    # 需要补齐的字节数
        if sup_len > 0:
            name = f"{' '*(sup_len//2)}{people[i]['Name']}{' '*(sup_len - sup_len//2)}"
        else:
            name = f"{people[i]['Name']}"[0:10]
        info = f"|{n: ^4}|{name}|{people[i]['NameId']: ^22}|{people[i]['DATETIME'].strftime('%Y-%m-%d %H:%M:%S'): ^25}|\n"
        text.insert('end', info)
        text.insert('end', '—' * 33 + "\n")
    text.config(state='disabled')
    SHOW = True
    # return people


def sched_func():
    if SHOW:
        show_people_info()
        print(datetime.now(), ' 刷新成功。')
    else:
        print('数据库正在读取中，忽略此次刷新。')


# 点击按钮下载
def download():
    dowm_date = e.get()
    # print(dowm_date)
    try:
        dowm_date_list = [int(i) for i in dowm_date.split("-")]
        # print(dowm_date_list)
        if len(dowm_date_list) != 3:
            e.delete(0, "end")
            e.insert(0, "格式有误")
        elif not(2050 >= dowm_date_list[0] >= 2021):
            e.delete(0, "end")
            e.insert(0, "年份有误")
        elif not(12 >= dowm_date_list[1] >= 1):
            e.delete(0, "end")
            e.insert(0, "月份有误")
        elif not (31 >= dowm_date_list[2] >= 1):
            e.delete(0, "end")
            e.insert(0, "日期有误")
        else:
            query_str = "-".join([str(i) for i in dowm_date_list])
            print(query_str)
            s_in = select(in_table).where(in_table.c.DATETIME.contains(query_str))
            res_in = con.execute(s_in).fetchall()
            s_out = select(out_table).where(out_table.c.DATETIME.contains(query_str))
            res_out = con.execute(s_out).fetchall()
            write_excel(query_str, res_in, res_out)

    except Exception as error:
        print(error)
        # e.delete(0, "end")
        # e.insert(0, "导出失败。")
        messagebox.showerror(message="导出失败！")
    # s = select(in_table).where(in_table.c.DATETIME.contains("2021-4-8"))
    # res = con.execute(s).fetchall()
    # print(len(res))
    # for i in res:
    #     print(i)


# 写入到excel
# 参数 d为查询日期，i 为进的记录，o 为出的记录。
def write_excel(d, i, o):
    fileName = filedialog.asksaveasfilename(title='请选择需要保存的文件夹',
                                            # initialdir=r'E:\360',
                                            initialfile=f'{d}.xlsx')
    new_excle = ex.Workbook()  # 新建一个工作簿。
    in_sheet = new_excle.active  # 把当前活动的工作表实例化给对象new_sheet。
    in_sheet.title = '进记录'
    in_sheet.cell(1, 1).value = '姓名'
    in_sheet.cell(1, 2).value = '卡号'
    in_sheet.cell(1, 3).value = '进门时间'
    in_sheet.cell(1, 4).value = '是否已出门'
    for row, item in enumerate(i, start=2):
        in_sheet.cell(row, 1).value = item['Name']
        in_sheet.cell(row, 2).value = item['NameId']
        in_sheet.cell(row, 3).value = item['DATETIME']
        in_sheet.cell(row, 4).value = "已出" if item['STATE'] == 1 else "未出"

    out_sheet = new_excle.create_sheet('出记录', 1)
    out_sheet.cell(1, 1).value = '姓名'
    out_sheet.cell(1, 2).value = '卡号'
    out_sheet.cell(1, 3).value = '出门时间'
    for row, item in enumerate(o, start=2):
        out_sheet.cell(row, 1).value = item['Name']
        out_sheet.cell(row, 2).value = item['NameId']
        out_sheet.cell(row, 3).value = item['DATETIME']

    # print(fileName)
    if fileName:
        new_excle.save(fileName)
        print('写入excel结束')
        messagebox.showinfo(message=f"已成功保存在: {fileName}")
    else:
        pass

# # 获取字段名
# key = in_table.c.keys()
# print(key)

# 获取主键字段
# print(in_table.primary_key)

# # 利用insert()方法插入数据，方法 1
# stmt = insert(in_table).values(Name='Lucy', NameId="222222", DATETIME=datetime.now())
# # print(stmt)
# session.execute(stmt)
# session.commit()

# # 利用insert()方法插入数据，方法 2，可以一次性插入多行。
# session.execute(insert(in_table), [{"Name":'Lili', "NameId":"33333", "DATETIME":datetime.now()}])
# session.commit()


if __name__ == '__main__':
    root = tk.Tk()
    root.title('人员统计')

    "获取电脑屏幕尺寸大小"
    window_x = root.winfo_screenwidth()
    window_y = root.winfo_screenheight()

    "设置窗口大小参数"
    WIDTH = 504
    HEIGHT = 610

    "获取窗口左上角的X, Y坐标，用来设置窗口的放置位置为屏幕的中间。"
    x = (window_x - WIDTH) / 2
    y = (window_y - HEIGHT) / 2
    root.geometry(f'{WIDTH}x{HEIGHT}+{int(x)}+{int(y)}')
    root.resizable(width=False, height=False)

    # 划分窗口
    top_frame = tk.Frame(root)
    top_frame.pack(side='top', pady=10, fill='x')

    mid_frame = tk.Frame(root)
    mid_frame.pack(side='top', pady=5, fill='both')

    v = tk.StringVar()
    label_1 = tk.Label(top_frame, textvariable=v,            # text为显示的文本内容
                                  font=("", 12),             # 设置字体为“华文行楷”，大小为20
                                  relief="ridge",
                                  width=20, height=2)       # width为标签的宽，height为高
    label_1.pack(side='left', anchor="w", padx=30)

    button_1 = tk.Button(top_frame, text=" 刷 新 ", command=sched_func)
    button_1.pack(side='left')

    e = tk.Entry(top_frame, width=12)
    e.pack(side="left", padx=30)
    e.insert(0, datetime.now().date())

    button_2 = tk.Button(top_frame, text=" 导出 ", command=download)
    button_2.pack(side='left')

    Sl = tk.Scrollbar(mid_frame)                # 建立 滚动条部件。
    Sl.config(orient='vertical')                # 设置滚动条显示方向：vertical=垂直，horizontal=水平。
    Sl.pack(side='right', padx=5, fill='y', expand=1)

    tk.Frame(mid_frame, width=10).pack(side='left')   # 只用来占位，使左侧显示空白。

    text = tk.Text(mid_frame, height=40)
    text.pack(side='left', pady=10, fill='x')

    text.config(yscrollcommand=Sl.set)
    Sl.config(command=text.yview)

    SHOW = True              # 可以读取数据库的信号
    show_people_info()       # 显示在房间内的人员。
    # print(people)

    # 创建刷新任务
    sched = BackgroundScheduler()
    sched.add_job(sched_func, trigger='interval', seconds=5)
    sched.start()

    tk.mainloop()

